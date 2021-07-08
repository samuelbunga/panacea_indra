import os
import re
import sys
import csv
import json
import tqdm
import pyobo
import obonet
import random
import pickle
import logging
import graphviz
import datetime
import openpyxl
import networkx
import itertools
import numpy as np
import pandas as pd
import enzyme_client
from pathlib import Path
from matplotlib import rc
from bioinfokit import visuz
from graphviz import Digraph
from indra.sources import tas
import matplotlib.pyplot as plt
from indra.util import batch_iter
from collections import OrderedDict
from collections import defaultdict
import matplotlib.colors as mcolors
from indra.statements import Complex
from scipy.stats import fisher_exact
from indra.sources import indra_db_rest
import indra.tools.assemble_corpus as ac
from indra.literature import pubmed_client
from indra.assemblers.cx import hub_layout
from indra.ontology.bio import bio_ontology
from indra.databases.uniprot_client import um
from indra.assemblers.html import HtmlAssembler
from indra.statements.agent import default_ns_order
from indra.sources.omnipath import process_from_web
from indra.assemblers.cx.assembler import CxAssembler
from indra.databases import uniprot_client, hgnc_client
from indra_db.client.principal.curation import get_curations
from indra.databases.hgnc_client import get_hgnc_from_mouse, get_hgnc_name


logger = logging.getLogger('receptor_ligand_interactions')

mouse_gene_name_to_mgi = {v: um.uniprot_mgi.get(k)
                          for k, v in um.uniprot_gene_name.items()
                          if k in um.uniprot_mgi}

up_hgnc = {v: k for k, v in um.uniprot_gene_name.items()
           if k in um.uniprot_hgnc}


HERE = os.path.dirname(os.path.abspath(__file__))
INPUT = os.path.join(HERE, os.pardir, 'input')
OUTPUT = os.path.join(HERE, os.pardir, 'output')
INDRA_DB_PKL = os.path.join(INPUT, 'db_dump_df.pkl')
GO_ANNOTATIONS = os.path.join(INPUT, 'goa_human.gaf')
ION_CHANNELS = os.path.join(INPUT, 'ion_channels.txt')
SURFACE_PROTEINS_WB = os.path.join(INPUT, 'Surface Proteins.xlsx')
wd = os.path.join(HERE, os.pardir)
	

db_curations = get_curations()



logger = logging.getLogger('receptor_ligand_interactions')

mouse_gene_name_to_mgi = {v: um.uniprot_mgi.get(k)
                          for k, v in um.uniprot_gene_name.items()
                          if k in um.uniprot_mgi}

up_hgnc = {v: k for k, v in um.uniprot_gene_name.items()
           if k in um.uniprot_hgnc}


db_curations = get_curations()


def _load_goa_gaf():
    """Load the gene/GO annotations as a pandas data frame."""
    # goa_ec = {'EXP', 'IDA', 'IPI', 'IMP', 'IGI', 'IEP', 'HTP', 'HDA', 'HMP',
    #          'HGI', 'HEP', 'IBA', 'IBD'}
    goa = pd.read_csv(GO_ANNOTATIONS, sep='\t',
                      skiprows=23, dtype=str,
                      header=None,
                      names=['DB',
                             'DB_ID',
                             'DB_Symbol',
                             'Qualifier',
                             'GO_ID',
                             'DB_Reference',
                             'Evidence_Code',
                             'With_From',
                             'Aspect',
                             'DB_Object_Name',
                             'DB_Object_Synonym',
                             'DB_Object_Type',
                             'Taxon',
                             'Date',
                             'Assigned',
                             'Annotation_Extension',
                             'Gene_Product_Form_ID'])
    goa = goa.sort_values(by=['DB_ID', 'GO_ID'])
    # Filter out all "NOT" negative evidences
    goa['Qualifier'].fillna('', inplace=True)
    goa = goa[~goa['Qualifier'].str.startswith('NOT')]
    # Filter to rows with evidence code corresponding to experimental
    # evidence
    # goa = goa[goa['Evidence_Code'].isin(goa_ec)]
    return goa


goa = _load_goa_gaf()


def load_indra_df(fname):
    """Return an INDRA Statement data frame from a pickle file."""
    logger.info('Loading INDRA DB dataframe')
    with open(fname, 'rb') as fh:
        df = pickle.load(fh)
    logger.info('Loaded %d rows from %s' % (len(df), fname))
    return df

# Load the INDRA DB DF
indra_df = load_indra_df(INDRA_DB_PKL)


def get_hashes_by_gene_pair(df, ligand_genes, receptor_genes):
    hashes_by_gene_pair = defaultdict(set)

    for a, b, hs in zip(df.agA_name, df.agB_name, df.stmt_hash):
        if a in ligand_genes and b in receptor_genes:
            hashes_by_gene_pair[(a, b)].add(hs)
    return hashes_by_gene_pair


def get_nature_hashes(df, nature_df):
    hashes_by_gene_pair = defaultdict(set)
    ligand_genes = list(nature_df.ligands)
    receptor_genes = list(nature_df.receptors)
    
    for a, b, hs in zip(df.agA_name, df.agB_name, df.stmt_hash):
        if a in ligand_genes:
            if b in receptor_genes[ligand_genes.index(a)]:
                hashes_by_gene_pair[(a, b)].add(hs)
    return hashes_by_gene_pair


def download_statements(hashes):
    """Download the INDRA Statements corresponding to a set of hashes.
    """
    stmts_by_hash = {}
    for group in tqdm.tqdm(batch_iter(hashes, 200), total=int(len(hashes) / 200)):
        idbp = indra_db_rest.get_statements_by_hash(list(group),
                                                    ev_limit=10)
        for stmt in idbp.statements:
            stmts_by_hash[stmt.get_hash()] = stmt
    return stmts_by_hash


def get_genes_for_go_ids(go_ids):
    """Return genes that are annotated with a given go ID or its children."""
    df = goa[goa['GO_ID'].isin(set(go_ids))]
    up_ids = sorted(list(set(df['DB_ID'])))
    gene_names = [uniprot_client.get_gene_name(up_id) for up_id in up_ids]
    gene_names = {g for g in gene_names if g}
    return gene_names





def read_workbook(workbook):
    """ This function takes Excel workbook as an input and
    returns ligand and receptor gene list respectively.
    Input: Excel workbook with single(2 columns) or two sheets
    Condition: considers first column/sheet as ligand genes and second
    column/shet as receptor genes
    """
    ligands_sheet = 'updated list of ligands '
    receptors_sheet = 'RPKM > 1.5 cfiber'
    wb = openpyxl.load_workbook(workbook)
    ligands = fix_dates(set([row[0].value for row in wb[ligands_sheet]][1:]))
    receptors = fix_dates(set([row[0].value
                               for row in wb[receptors_sheet]][1:]))
    return ligands, receptors


def read_gene_list(infile, mode):
    gene_list = []
    try:
        with open(infile, mode) as FH:
            for eachGene in FH:
                gene_list.append(eachGene.strip("\n"))
        return gene_list

    except FileNotFoundError:
        sys.exit("Given file doesn't exist")


def filter_complex_statements(stmts, ligands, receptors):
    filtered_stmts = []
    for stmt in stmts:
        if isinstance(stmt, Complex):
            if len(stmt.members) <=2:
                if (any(a.name in ligands for a in stmt.members) 
                    and any(a.name in receptors for a in stmt.members)):
                    filtered_stmts.append(stmt)
        else:
            filtered_stmts.append(stmt)


    return filtered_stmts


def filter_op_stmts(op_stmts, lg, rg):
    """ Filter out the statements which are not ligand and receptor """
    logger.info(f'Filtering {len(op_stmts)} to ligand-receptor interactions')
    filtered_stmts = [stmt for stmt in op_stmts if
                      (any(a.name in lg for a in stmt.agent_list())
                       and any(a.name in rg for a in stmt.agent_list()))]
    logger.info(f'{len(filtered_stmts)} left after filter')
    return filtered_stmts


def html_assembler(indra_stmts, fname):
    """Assemble INDRA statements into a HTML report"""
    html_assembler = HtmlAssembler(indra_stmts,
                                   db_rest_url='https://db.indra.bio')
    assembled_html_report = html_assembler.make_model(no_redundancy=True)
    html_assembler.save_model(fname)
    return assembled_html_report


def cx_assembler(indra_stmts, fname):
    """Assemble INDRA statements into a CX report"""
    cx_assembler = CxAssembler(indra_stmts)
    assembled_cx_report = cx_assembler.make_model()
    cx_assembler.save_model(fname)
    ndex_network_id = cx_assembler.upload_model(ndex_cred=None,
                                                private=True, style='default')
    return assembled_cx_report, ndex_network_id



def get_receptor_by_ligands(receptors_in_data, ligands_in_data, stmts):
    receptor_by_ligands = defaultdict(set)
    for stmt in stmts:
        agent_names = {agent.name for agent in stmt.agent_list()}
        receptors = agent_names & receptors_in_data
        ligands = agent_names & ligands_in_data
        for receptor in receptors:
            receptor_by_ligands[receptor] |= ligands
    return dict(receptor_by_ligands)


def filter_out_medscan(stmts):
    logger.info('Filtering out medscan evidence on %d statements' % len(stmts))
    new_stmts = []
    for stmt in stmts:
        new_evidence = [e for e in stmt.evidence if e.source_api != 'medscan']
        if not new_evidence:
            continue
        stmt.evidence = new_evidence
        if not stmt.evidence:
            continue
        new_stmts.append(stmt)
    logger.info('%d statements after filter' % len(new_stmts))
    return new_stmts


def filter_db_only(stmts):
    new_stmts = []
    for stmt in stmts:
        sources = {ev.source_api for ev in stmt.evidence}
        if sources <= {'reach', 'sparser', 'trips', 'rlimsp', 'medscan', 'eidos'}:
            continue
        new_stmts.append(stmt)
    return new_stmts


def filter_incorrect_curations(stmts):
    # Filter incorrect curations
    indra_op_filtered = ac.filter_by_curation(stmts,
                                              curations=db_curations)
    return indra_op_filtered


def mgi_to_hgnc_name(gene_list):
    """Convert given mouse gene symbols to HGNC equivalent symbols"""
    filtered_mgi = {mouse_gene_name_to_mgi[gene] for gene in gene_list
                    if gene in mouse_gene_name_to_mgi}
    hgnc_gene_set = set()
    for mgi_id in filtered_mgi:
        hgnc_id = get_hgnc_from_mouse(mgi_id)
        hgnc_gene_set.add(get_hgnc_name(hgnc_id))
    return hgnc_gene_set


def process_df(workbook):
    wb = openpyxl.load_workbook(workbook)
    df = {
    'ligands': [row[1].value for row in wb['All.Pairs']][1:], 
    'receptors': [row[3].value for row in wb['All.Pairs']][1:]
    }
    lg_rg = pd.DataFrame(df)
    return lg_rg


def expand_with_child_go_terms(terms):
    all_terms = set(terms)
    for term in terms:
        child_terms = bio_ontology.get_children('GO', term)
        all_terms |= {c[1] for c in child_terms}
    return all_terms


def get_receptors():
    receptor_terms = ['signaling receptor activity']
    receptor_go_ids = [bio_ontology.get_id_from_name('GO', term)[1]
                       for term in receptor_terms]
    receptor_go_ids = expand_with_child_go_terms(receptor_go_ids)
    # Filtering out the nuclear receptors from the receptor list
    receptor_go_ids = {r for r in receptor_go_ids if 'receptor' in
                       bio_ontology.get_name('GO', r)} - \
                      expand_with_child_go_terms(['GO:0004879'])
    receptor_genes_go = get_genes_for_go_ids(receptor_go_ids)
    receptor_genes_go -= {'NR2C2', 'EGF'}
    # Add ION channels to the receptor list
    ion_channels = set()
    with open(ION_CHANNELS, 'r') as fh:
        for line in fh:
            ion_channels.add(line.strip())
    receptor_genes_go |= ion_channels
    return receptor_genes_go


def get_ligands():
    # Read and extract cell surface proteins from CSPA DB
    wb = openpyxl.load_workbook(SURFACE_PROTEINS_WB)
    surface_protein_set = set(row[4].value for row in wb['Sheet 1']
                              if row[6].value == 'yes')
    logger.info('Got %d surface proteins from spreadsheet' %
                len(surface_protein_set))
    ligand_terms = ['cytokine activity', 'hormone activity',
                    'growth factor activity',
                    'extracellular matrix structural constituent']
    # Getting GO id's for ligands and receptors by using
    # GO terms
    ligand_go_ids = [bio_ontology.get_id_from_name('GO', term)[1]
                     for term in ligand_terms]
    ligand_go_ids = expand_with_child_go_terms(ligand_go_ids)

    # Converting GO id's to gene symbols
    ligand_genes_go = get_genes_for_go_ids(ligand_go_ids)
    # Remove one more nuclear receptor
    #manual_ligands = {'THBS1'}
    manual_ligands = set()
    return surface_protein_set | ligand_genes_go | manual_ligands


def process_nature_paper():
    nature_interactions = process_df(os.path.join(INPUT, 'ncomms8866-s3.xlsx'))
    custom_interactome = defaultdict(set)
    
    for r,c in nature_interactions.iterrows():
        custom_interactome[(c[0])].add(c[1])
        
    nature_dataframe = []
    count=0

    for r,c in nature_interactions.iterrows():
        count+=1
        if c[0] in up_hgnc and c[1] in up_hgnc:
            nature_dataframe.append(
                {
                    'id_cp_interaction':'NATURE-'+str(count),
                    'partner_a': up_hgnc[c[0]],
                    'partner_b': up_hgnc[c[1]],
                    'source':'NATURE'
                }
            )

    nature_dataframe = pd.DataFrame(nature_dataframe)
    nature_dataframe.to_csv(os.path.join(wd, 'output/nature_uniprot.csv'), 
                            sep=",", index=0)
    return nature_interactions
    
    
def merge_interactions(interactions, genes_file, uniprot_file):
    
    df = [{'partner_a':i.split('_')[0],
           'partner_b':i.split('_')[1]} 
          for i in interactions]

    interactions_hgnc = pd.DataFrame(df)
    interactions_hgnc.to_csv(os.path.join(wd, 'output', genes_file), 
                             sep=",", index=0)
    

    cellphonedb_df = []
    count=0
    for r,c in interactions_hgnc.iterrows():
        count+=1
        if c[0] in up_hgnc and c[1] in up_hgnc:
            cellphonedb_df.append(
                {
                    'id_cp_interaction':'Woolf-'+str(count),
                    'partner_a': up_hgnc[c[0]],
                    'partner_b': up_hgnc[c[1]],
                    'source':'INDRA'
                }
            )

    cellphonedb_df = pd.DataFrame(cellphonedb_df)
    cellphonedb_df.to_csv(os.path.join(wd, 'output', uniprot_file), 
                          sep=",", index=0)

if __name__ == '__main__':

	receptor_genes_go = get_receptors()
	# remove all the receptors from the surface_protein_set
	full_ligand_set = get_ligands() - receptor_genes_go

	# Now get INDRA DB Statements for the receptor-ligand pairs
	hashes_by_gene_pair = get_hashes_by_gene_pair(indra_df, full_ligand_set,
	                                              receptor_genes_go)

	# get the union of all the statement hashes 
	all_hashes = set.union(*hashes_by_gene_pair.values())
	# Download the statements by hashes
	stmts_by_hash = download_statements(all_hashes)
	# get only the list of all the available statemtns
	indra_db_stmts = list(stmts_by_hash.values())

	# Filtering out the indirect INDRA statements
	indra_db_stmts = ac.filter_direct(indra_db_stmts)
	        

	# Fetch omnipath database biomolecular interactions and
	# process them into INDRA statements
	op = process_from_web()

	# Filter statements which are not ligands/receptors from
	# OmniPath database
	op_filtered = filter_op_stmts(op.statements, full_ligand_set,
	                              receptor_genes_go)
	op_filtered = ac.filter_direct(op_filtered)


	op_filtered = ac.filter_by_curation(op_filtered,
	                                    curations=db_curations)


	# Merge omnipath/INDRA statements and run assembly
	indra_op_stmts = ac.run_preassembly(indra_db_stmts + op_filtered,
	                                    run_refinement=False)
	# Filter incorrect curations        
	indra_op_filtered = filter_incorrect_curations(indra_op_stmts)

	# Filter complex statements
	indra_op_filtered = filter_complex_statements(indra_op_filtered,
	                                              full_ligand_set,
	                                              receptor_genes_go)

	# We do this again because when removing complex members, we
	# end up with more duplicates
	indra_op_filtered = ac.run_preassembly(indra_op_filtered,
	                                       run_refinement=False)
	    

	# Filter complex OP statements
	op_filtered = filter_complex_statements(op_filtered,
	                                        full_ligand_set,
	                                        receptor_genes_go)
	op_filtered = ac.run_preassembly(op_filtered, run_refinement=False)

	# Filtering indra_db stmts
	indra_db_stmts = ac.filter_by_curation(indra_db_stmts,
	                                       curations=db_curations)

	indra_db_stmts = filter_complex_statements(indra_db_stmts,
	                                           full_ligand_set,
	                                           receptor_genes_go)

	indra_db_stmts = ac.run_preassembly(indra_db_stmts, run_refinement=False)



	# get receptor by ligands for OP
	op_receptor_by_ligands = get_receptor_by_ligands(receptor_genes_go, 
	                                                 full_ligand_set, 
	                                                 op_filtered)

	# get receptor by ligands for indra_op
	indra_op_receptor_by_ligands = get_receptor_by_ligands(receptor_genes_go, 
	                                                       full_ligand_set, 
	                                                       indra_op_filtered)

	# get receptor by ligands for indra db
	indra_db_receptor_by_ligands = get_receptor_by_ligands(receptor_genes_go, 
	                                                       full_ligand_set, 
	                                                       indra_db_stmts)


	# Assemble the statements into HTML formatted report and save into a file
	op_html_report = \
	    html_assembler(
	        op_filtered,
	        fname=os.path.join(wd, 'output', 'op_interactions.html'))

	indra_op_html_report = \
	        html_assembler(
	            indra_op_filtered,
	            fname=os.path.join(wd, 'output', 'indra_op_interactions.html'))



	nature_interactions = process_nature_paper()

	# Make nature interactions dataframe
	nature_interactions_df = []
	for r,c in nature_interactions.iterrows():
	    nature_interactions_df.append(
	    {
	        'ligands': c[0],
	        'receptors': c[1],
	        'interactions': c[0]+'_'+c[1]

	    }
	)

	nature_interactions_df = pd.DataFrame(nature_interactions_df)


	# Make OP interactions dataframe
	op_interactions = []
	for receptors, ligands in op_receptor_by_ligands.items():
	    for lg in ligands:
	        op_interactions.append(
	        {
	            'ligands':lg,
	            'receptors': receptors,
	            'interactions': lg+'_'+receptors
	        }
	    )
	op_df = pd.DataFrame(op_interactions)


	# Make INDRA_OP interactions dataframe
	indra_op_interactions = []
	for receptors, ligands in indra_op_receptor_by_ligands.items():
	    for lg in ligands:
	        indra_op_interactions.append(
	        {
	            'ligands':lg,
	            'receptors': receptors,
	            'interactions': lg+'_'+receptors
	        }
	    )
	indra_op_df = pd.DataFrame(indra_op_interactions)


	# Make INDRA_DB interactions dataframe
	indra_db_interactions = []
	for receptors, ligands in indra_db_receptor_by_ligands.items():
	    for lg in ligands:
	        indra_db_interactions.append(
	        {
	            'ligands':lg,
	            'receptors': receptors,
	            'interactions': lg+'_'+receptors
	        }
	    )
	indra_db_df = pd.DataFrame(indra_db_interactions)


	unique_op_interactions = set(op_df.interactions)
	unique_nature_interactions = set(nature_interactions_df.interactions)
	unique_indra_op_interactions = set(indra_op_df.interactions)
	unique_indra_db_interactions = set(indra_db_df.interactions)
	unique_indra_db_only = unique_indra_db_interactions - (unique_op_interactions | unique_nature_interactions)

	# Nature and OP interactions only
	unique_op_nature_interactions = unique_op_interactions | unique_nature_interactions

	logger.info('Total OP interactions: %d'% len(unique_op_interactions))
	logger.info('Total Nature interactions: %d'% len(unique_nature_interactions))
	logger.info('Total OP and Nature interactions: %d'% len(unique_op_nature_interactions))

	merge_interactions(unique_op_nature_interactions, 'op_nature_genes.csv', 'op_nature_uniprot.csv')


	# Nature and INDRA_OP interactions only
	unique_indra_op_nature_interactions = unique_indra_op_interactions | unique_nature_interactions

	logger.info('Total INDRA_OP interactions: %d'% len(unique_indra_op_interactions))
	logger.info('Total Nature interactions: %d'% len(unique_nature_interactions))
	logger.info('Total INDRA_OP and Nature interactions: %d'% len(unique_indra_op_nature_interactions))

	merge_interactions(unique_indra_op_nature_interactions, 'indra_op_nature_genes.csv', 'indra_op_nature_uniprot.csv')

	filtered_indra_db_stmts = []
	f = [s.split('_') for s in unique_indra_db_only]

	for stmts in indra_db_stmts:
	    agents_list = [agents.name for agents in stmts.agent_list()]
	    if agents_list in f:
	        filtered_indra_db_stmts.append(stmts)
	    
	    
	indra_db_only_html_report = \
	        html_assembler(
	            filtered_indra_db_stmts,
	            fname=os.path.join(wd, 'output', 'indra_db_only_interactions.html'))


	######## Checking for OP specific and common interaction b/w Nature and OP
	#common_op_nature_interaction = set(op_df.interactions) & set(nature_df.interactions)
	#common_op_nature_interaction = pd.DataFrame([{'interactions':i} for i in common_op_nature_interaction])
	#common_op_nature_interaction.to_csv(os.path.join(wd, 'output/common_op_nature_interaction.csv'), 
	#                 sep=",", index=0)

	#op_specific = set(op_df.interactions) - set(nature_df.interactions)
	#op_specific = pd.DataFrame([{'interactions':i} for i in op_specific])
	#op_specific.to_csv(os.path.join(wd, 'output/op_specific.csv'), 
	#                 sep=",", index=0)